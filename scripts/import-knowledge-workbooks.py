#!/usr/bin/env python3
"""Normalize the approved news, report-date and FID workbooks into versioned JSON."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

SYNCED_ON = "2026-08-07"
NEWS_CATEGORIES = (
    "公司动态",
    "项目进展",
    "产量与运营",
    "财务与交易",
    "能源转型",
    "政策与市场",
)

CATEGORY_RULES = {
    "能源转型": (
        "renewable", "wind", "solar", "hydrogen", "ammonia", "carbon capture", "ccs", "ccus",
        "decarbon", "clean energy", "energy transition", "biofuel", "battery", "绿色", "风电", "光伏",
        "氢", "氨", "碳捕集", "碳中和", "能源转型", "清洁能源", "生物燃料",
    ),
    "财务与交易": (
        "earnings", "financial result", "profit", "revenue", "merger", "acquisition", "acquire",
        "divest", "sale", "stake", "investment", "financing", "funding", "ipo", "dividend", "bond",
        "业绩", "利润", "营收", "财务", "并购", "合并", "收购", "出售", "股权", "投资", "融资",
        "上市", "股息", "债券",
    ),
    "项目进展": (
        "final investment decision", " fid ", "contract", "award", "construction", "build", "project",
        "development", "commission", "start-up", "startup", "launch", "approval", "permit", "tender",
        "epc", "farm-in", "farmout", "farm-out", "项目", "最终投资决定", "合同", "授标", "开工",
        "建设", "投产", "启动", "获批", "许可", "招标", "开发",
    ),
    "产量与运营": (
        "production", "output", "drilling", "well", "discovery", "exploration", "operation", "shutdown",
        "outage", "maintenance", "cargo", "export", "import", "reserve", "产量", "生产", "钻井", "油井",
        "发现", "勘探", "运营", "停产", "检修", "货物", "出口", "进口", "储量",
    ),
    "政策与市场": (
        "policy", "regulation", "sanction", "tariff", "price", "market", "forecast", "outlook", "demand",
        "supply", "trade", "opec", "iea", "政策", "监管", "制裁", "关税", "价格", "市场", "预测",
        "展望", "需求", "供应", "贸易",
    ),
}


def arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--news-recent", required=True, type=Path)
    parser.add_argument("--news-market", required=True, type=Path)
    parser.add_argument("--reports", required=True, type=Path)
    parser.add_argument("--fid", required=True, type=Path)
    parser.add_argument("--repository-root", type=Path, default=Path(__file__).resolve().parents[1])
    return parser.parse_args()


def normalize_name(value: object) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", str(value or "").strip().casefold())


def slugify(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode().lower()
    text = text.replace("&", " and ")
    return re.sub(r"^-+|-+$", "", re.sub(r"[^a-z0-9]+", "-", text))


def clean(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def iso_date(value: object, workbook_epoch: datetime) -> str | None:
    if value in (None, "", "-"):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return from_excel(value, workbook_epoch).date().isoformat()
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"Unsupported date value: {value!r}")


def canonical_url(value: object) -> str | None:
    raw = clean(value)
    if not raw:
        return None
    if "://" not in raw:
        raw = f"https://{raw.lstrip('/')}"
    parts = urlsplit(raw)
    if parts.scheme not in {"http", "https"} or not parts.netloc:
        raise ValueError(f"Invalid news URL: {raw}")
    return urlunsplit(("https", parts.netloc.lower(), parts.path.rstrip("/"), "", ""))


def workbook_rows(path: Path) -> tuple[list[str], list[dict[str, object]], datetime]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
    while headers and not headers[-1]:
        headers.pop()
    return headers, [dict(zip(headers, values[:len(headers)])) for values in iterator], workbook.epoch


def company_aliases(root: Path) -> tuple[dict[str, str], set[str]]:
    profiles = json.loads((root / "company-text-dashboard/data/company-data.json").read_text())
    operators = json.loads((root / "maps/operators.json").read_text())["operators"]
    inventory = json.loads((root / "data/company-demo-inventory.json").read_text())
    overrides = json.loads((root / "data/company-alias-overrides.json").read_text())
    featured = {normalize_name(item["displayName"]): item["slug"] for item in inventory["companies"]}
    operator_by_name = {
        normalize_name(name): operator["slug"]
        for operator in operators
        for name in [operator["name"], *operator.get("aliases", [])]
    }
    aliases = dict(operator_by_name)
    slugs = set()
    for profile in profiles:
        key = normalize_name(profile["公司名称"])
        slug = featured.get(key) or operator_by_name.get(key) or slugify(profile["公司名称"])
        aliases[key] = slug
        slugs.add(slug)
    for name, slug in overrides.items():
        if slug not in slugs:
            raise ValueError(f"Alias {name!r} references unknown company slug {slug!r}")
        aliases[normalize_name(name)] = slug
    if len(slugs) != 126:
        raise ValueError(f"Expected 126 company slugs, found {len(slugs)}")
    return aliases, slugs


def news_category(*values: object) -> str:
    text = f" {' '.join(str(value or '') for value in values).casefold()} "
    for category, keywords in CATEGORY_RULES.items():
        if any(keyword in text for keyword in keywords):
            return category
    return "公司动态"


def merge_text(current: str | None, incoming: object) -> str | None:
    candidate = clean(incoming)
    if not candidate:
        return current
    return candidate if not current or len(candidate) > len(current) else current


def build_news(paths: list[Path], aliases: dict[str, str]) -> dict[str, object]:
    stories: dict[str, dict[str, object]] = {}
    source_rows = 0
    for path in paths:
        headers, rows, epoch = workbook_rows(path)
        required = {"中文标题", "英文标题", "中文摘要", "英文摘要", "发布日期", "新闻链接", "新闻来源", "地区", "公司"}
        if set(headers) != required:
            raise ValueError(f"Unexpected news columns in {path.name}: {headers}")
        for row_number, row in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in row.values()):
                continue
            source_rows += 1
            published_on = iso_date(row.get("发布日期"), epoch)
            source_url = canonical_url(row.get("新闻链接"))
            chinese_title = clean(row.get("中文标题"))
            english_title = clean(row.get("英文标题"))
            if not chinese_title or not english_title or not published_on:
                raise ValueError(
                    f"News row {row_number} is missing a bilingual title or date in {path.name}: "
                    f"zh={chinese_title!r}, en={english_title!r}, date={published_on!r}"
                )
            identity = source_url or f"{normalize_name(chinese_title)}|{published_on}"
            story_id = f"news-{hashlib.sha256(identity.encode()).hexdigest()[:24]}"
            story = stories.setdefault(story_id, {
                "id": story_id,
                "title": chinese_title,
                "subtitle": english_title,
                "summary": clean(row.get("中文摘要")),
                "summaryEn": clean(row.get("英文摘要")),
                "publisher": clean(row.get("新闻来源")) or "待核实",
                "publishedOn": published_on,
                "sourceUrl": source_url,
                "region": clean(row.get("地区")) or "全球/跨区域",
                "category": "公司动态",
                "companyNames": set(),
                "companySlugs": set(),
                "sourceFiles": set(),
            })
            story["title"] = merge_text(story["title"], chinese_title)
            story["subtitle"] = merge_text(story["subtitle"], english_title)
            story["summary"] = merge_text(story["summary"], row.get("中文摘要"))
            story["summaryEn"] = merge_text(story["summaryEn"], row.get("英文摘要"))
            story["sourceFiles"].add(path.name)
            company_name = clean(row.get("公司"))
            if company_name:
                story["companyNames"].add(company_name)
                company_slug = aliases.get(normalize_name(company_name))
                if company_slug:
                    story["companySlugs"].add(company_slug)
    for story in stories.values():
        story["category"] = news_category(story["title"], story["subtitle"], story["summary"], story["summaryEn"])
        for key in ("companyNames", "companySlugs", "sourceFiles"):
            story[key] = sorted(story[key])
    ordered = sorted(stories.values(), key=lambda item: (item["publishedOn"], item["id"]), reverse=True)
    unmatched = Counter(
        name for story in ordered for name in story["companyNames"]
        if normalize_name(name) not in aliases
    )
    return {
        "syncedOn": SYNCED_ON,
        "sourceFiles": [path.name for path in paths],
        "sourceRows": source_rows,
        "categories": list(NEWS_CATEGORIES),
        "unmatchedCompanyNames": dict(sorted(unmatched.items())),
        "news": ordered,
    }


def build_fid(path: Path, aliases: dict[str, str]) -> dict[str, object]:
    headers, rows, _ = workbook_rows(path)
    required = ["data_id", "运营商", "项目", "批准年份", "资产", "油气田类型", "设施类别", "权益", "国家", "历史所属公司", "经济性（百万美元）"]
    if headers != required:
        raise ValueError(f"Unexpected FID columns: {headers}")
    groups: dict[tuple[object, ...], dict[str, object]] = {}
    unmatched = Counter()
    seen_ids = set()
    source_rows = 0
    for row in rows:
        source_id = clean(row.get("data_id"))
        operator = clean(row.get("运营商"))
        if not source_id or operator == "Total":
            continue
        source_rows += 1
        if source_id in seen_ids:
            raise ValueError(f"Duplicate FID data_id: {source_id}")
        seen_ids.add(source_id)
        company_slug = aliases.get(normalize_name(operator))
        economics = row.get("经济性（百万美元）")
        visible_values = {
            "operatorName": operator or "未提供",
            "companySlug": company_slug,
            "project": clean(row.get("项目")) or "未提供",
            "approvalYear": clean(row.get("批准年份")),
            "asset": clean(row.get("资产")) or "未提供",
            "fieldType": clean(row.get("油气田类型")) or "未提供",
            "facilityCategory": clean(row.get("设施类别")) or "未提供",
            "interests": clean(row.get("权益")) or "未提供",
            "country": clean(row.get("国家")) or "未提供",
        }
        key = tuple(visible_values.values())
        group = groups.setdefault(key, {
            **visible_values,
            "sourceIds": [],
            "economicsCandidates": [],
        })
        group["sourceIds"].append(source_id)
        group["economicsCandidates"].append({
            "historicalCompany": clean(row.get("历史所属公司")),
            "value": round(float(economics), 6) if economics is not None else None,
        })
    projects = []
    for group in groups.values():
        operator = group["operatorName"]
        company_slug = group["companySlug"]
        if not company_slug:
            unmatched[operator] += 1
        candidates = group.pop("economicsCandidates")
        preferred = [
            candidate["value"] for candidate in candidates
            if normalize_name(candidate["historicalCompany"]) == normalize_name(operator)
            or (company_slug and aliases.get(normalize_name(candidate["historicalCompany"])) == company_slug)
        ]
        distinct = {candidate["value"] for candidate in candidates}
        economics_value = preferred[0] if preferred else (next(iter(distinct)) if len(distinct) == 1 else None)
        source_ids = sorted(group["sourceIds"])
        projects.append({
            "id": source_ids[0],
            **group,
            "economicsUsdMillion": economics_value,
        })
    return {
        "syncedOn": SYNCED_ON,
        "sourceFile": path.name,
        "sourceRows": source_rows,
        "deduplication": "Operator + project + approval year + asset + field type + facility category + interests + country",
        "unmatchedOperators": dict(sorted(unmatched.items())),
        "projects": projects,
    }


def update_reports(path: Path, root: Path) -> dict[str, int]:
    headers, rows, epoch = workbook_rows(path)
    expected = ["data_id", "报告名", "资料类型", "报告类型", "关联公司", "发布机构", "行业", "发布日期"]
    if headers != expected:
        raise ValueError(f"Unexpected report columns: {headers}")
    catalog_path = root / "data/report-catalog.json"
    overrides_path = root / "data/report-date-overrides.json"
    catalog = json.loads(catalog_path.read_text())
    by_source_id = {str(report["sourceRecordId"]): report for report in catalog["reports"]}
    dates: dict[str, str] = {}
    updated = 0
    for row in rows:
        source_id = clean(row.get("data_id"))
        if source_id not in by_source_id:
            raise ValueError(f"Report source ID is not present in the catalog: {source_id}")
        published_on = iso_date(row.get("发布日期"), epoch)
        if not published_on:
            raise ValueError(f"Report publication date is missing: {source_id}")
        dates[source_id] = published_on
        if by_source_id[source_id]["publishedOn"] != published_on:
            by_source_id[source_id]["publishedOn"] = published_on
            updated += 1
    if len(rows) != 722:
        raise ValueError(f"Expected 722 report-date rows, found {len(rows)}")
    overrides_path.write_text(json.dumps({
        "syncedOn": SYNCED_ON,
        "sourceFile": path.name,
        "dates": dict(sorted(dates.items())),
    }, ensure_ascii=False, indent=2) + "\n")
    catalog["syncedOn"] = SYNCED_ON
    catalog_path.write_text(json.dumps(catalog, ensure_ascii=False, indent=2) + "\n")
    return {"rows": len(rows), "updatedDates": updated}


def main() -> None:
    args = arguments()
    root = args.repository_root.resolve()
    aliases, company_slugs = company_aliases(root)
    news = build_news([args.news_recent, args.news_market], aliases)
    fid = build_fid(args.fid, aliases)
    report_result = update_reports(args.reports, root)
    (root / "data/news-catalog.json").write_text(json.dumps(news, ensure_ascii=False, indent=2) + "\n")
    (root / "data/fid-projects.json").write_text(json.dumps(fid, ensure_ascii=False, indent=2) + "\n")
    summary = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "companies": len(company_slugs),
        "news": {
            "sourceRows": news["sourceRows"],
            "stories": len(news["news"]),
            "relationships": sum(len(item["companySlugs"]) for item in news["news"]),
            "unmatchedNames": len(news["unmatchedCompanyNames"]),
            "categories": dict(Counter(item["category"] for item in news["news"])),
        },
        "fid": {
            "sourceRows": fid["sourceRows"],
            "rows": len(fid["projects"]),
            "matchedRows": sum(item["companySlug"] is not None for item in fid["projects"]),
            "matchedCompanies": len({item["companySlug"] for item in fid["projects"] if item["companySlug"]}),
            "unmatchedOperators": len(fid["unmatchedOperators"]),
        },
        "reports": report_result,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
